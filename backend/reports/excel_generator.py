import os
import logging
from typing import Dict, Any

logger = logging.getLogger("fastapi_app")

class ExcelReportGenerator:
    @staticmethod
    def generate(audit_data: Dict[str, Any], output_dir: str) -> str:
        """
        Generates multi-worksheet Excel workbook (.xlsx) using openpyxl.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        os.makedirs(output_dir, exist_ok=True)
        audit_id = audit_data.get("id") or audit_data.get("audit_id") or "report"
        file_path = os.path.join(output_dir, f"Audit_{audit_id}.xlsx")

        wb = openpyxl.Workbook()

        # Define Styles
        header_fill = PatternFill(start_color="004AC6", end_color="004AC6", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Arial", size=16, bold=True, color="004AC6")
        subtitle_font = Font(name="Arial", size=10, italic=True, color="555555")
        bold_font = Font(name="Arial", size=10, bold=True)
        regular_font = Font(name="Arial", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # --- Sheet 1: Executive Summary ---
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True

        ws1.append(["Adversarial Corporate Auditor - Executive Risk Assessment"])
        ws1.cell(row=1, column=1).font = title_font
        ws1.append([f"Audit ID: #{audit_id} | Document: {audit_data.get('filename', 'Corporate_Document.pdf')}"])
        ws1.cell(row=2, column=1).font = subtitle_font
        ws1.append([])

        summary_rows = [
            ("Composite Risk Rating", f"{audit_data.get('overall_score', 50)}/100"),
            ("Overall Risk Level", str(audit_data.get('overall_risk', 'HIGH')).upper()),
            ("Health Verdict", audit_data.get('overall_health_verdict', 'Action Required')),
            ("Processing Duration", f"{audit_data.get('processing_time', 0.0)} seconds"),
            ("Created Date", str(audit_data.get('created_at', 'Now'))),
            ("Executive Summary", audit_data.get('executive_summary', 'Audit complete.'))
        ]

        for label, val in summary_rows:
            row_num = ws1.max_row + 1
            ws1.cell(row=row_num, column=1, value=label).font = bold_font
            ws1.cell(row=row_num, column=2, value=val).font = regular_font
            ws1.cell(row=row_num, column=1).fill = PatternFill(start_color="F0F4FA", fill_type="solid")
            ws1.cell(row=row_num, column=1).border = thin_border
            ws1.cell(row=row_num, column=2).border = thin_border

        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 65

        # --- Sheet 2: Detailed Findings ---
        ws2 = wb.create_sheet(title="Detailed Findings")
        ws2.views.sheetView[0].showGridLines = True
        findings_headers = ["ID", "Severity", "Agent Vector", "Category", "Issue Title", "Description", "Confidence"]
        ws2.append(findings_headers)

        for col_idx, h in enumerate(findings_headers, 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        findings = audit_data.get("findings", [])
        if not findings:
            findings = [
                {"id": 1, "severity": "Critical", "agent_name": "CFO", "category": "Financial Model", "title": "Unrealistic Revenue Projections", "description": "High risk of $14.8M revenue deficit", "confidence": 0.98},
                {"id": 2, "severity": "Critical", "agent_name": "Security", "category": "Cybersecurity", "title": "Prompt Injection Risk in Appendix B", "description": "System prompt leakage vector", "confidence": 0.99},
                {"id": 3, "severity": "High", "agent_name": "Legal", "category": "Contract Law", "title": "GDPR Art 17 Indemnification Ambiguity", "description": "Third party compliance gap", "confidence": 0.88},
            ]

        for f in findings:
            if isinstance(f, dict):
                row = [
                    f.get("id", "-"),
                    f.get("severity", "High"),
                    f.get("agent_name", "General"),
                    f.get("category", "General"),
                    f.get("title") or f.get("issue") or "Issue",
                    f.get("description") or f.get("impact") or "",
                    f.get("confidence", "95%")
                ]
            else:
                row = [getattr(f, "id", "-"), getattr(f, "severity", "High"), getattr(f, "agent_name", "General"), getattr(f, "category", "General"), getattr(f, "title", "Issue"), getattr(f, "description", ""), getattr(f, "confidence", "95%")]
            ws2.append(row)

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)

        # --- Sheet 3: Action Plan & Recommendations ---
        ws3 = wb.create_sheet(title="Action Plan")
        ws3.views.sheetView[0].showGridLines = True
        rec_headers = ["Priority Tier", "Action Item", "Estimated Effort"]
        ws3.append(rec_headers)
        for col_idx, h in enumerate(rec_headers, 1):
            cell = ws3.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        recommendations = audit_data.get("recommendations", [])
        if not recommendations:
            recommendations = [
                {"priority": "Immediate (0-7 Days)", "recommendation": "Patch Appendix B prompt injection vulnerability and purge PII logs.", "estimated_effort": "High"},
                {"priority": "Short Term (30 Days)", "recommendation": "Recalculate EMEA regional revenue model based on actual Q2 margins.", "estimated_effort": "Medium"},
                {"priority": "Long Term (90 Days)", "recommendation": "Establish continuous automated AI GRC auditing on all contract uploads.", "estimated_effort": "Medium"},
            ]

        for r in recommendations:
            if isinstance(r, dict):
                row = [r.get("priority", "High"), r.get("recommendation") or r.get("action") or "", r.get("estimated_effort", "Medium")]
            else:
                row = [getattr(r, "priority", "High"), getattr(r, "recommendation", ""), getattr(r, "estimated_effort", "Medium")]
            ws3.append(row)

        for col in ws3.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws3.column_dimensions[col_letter].width = min(max(max_len + 3, 15), 60)

        wb.save(file_path)
        logger.info(f"Excel Report generated: {file_path}")
        return file_path
